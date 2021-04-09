from openpyxl.styles import Color, PatternFill, Font, Border
from bs4 import BeautifulSoup
import openpyxl as xl
import requests
# import threaded
# import schedule
import datetime
import time
import random
import re
import os
import shutil
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.alert import Alert

from selenium.common.exceptions import TimeoutException

from selenium.webdriver.chrome.options import Options
# options = webdriver.ChromeOptions()


def GetProxies():    
    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    # driver = webdriver.Chrome(chrome_options=options, executable_path=r'C:\WebDrivers\chromedriver.exe')
    driver = webdriver.Chrome('chromedriver', options=options)
    driver.get("https://sslproxies.org/")
    driver.execute_script("return arguments[0].scrollIntoView(true);", WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, "//table[@class='table table-striped table-bordered dataTable']//th[contains(., 'IP Address')]"))))
    ips = [my_elem.get_attribute("innerHTML") for my_elem in WebDriverWait(driver, 5).until(EC.visibility_of_all_elements_located((By.XPATH, "//table[@class='table table-striped table-bordered dataTable']//tbody//tr[@role='row']/td[position() = 1]")))]
    ports = [my_elem.get_attribute("innerHTML") for my_elem in WebDriverWait(driver, 5).until(EC.visibility_of_all_elements_located((By.XPATH, "//table[@class='table table-striped table-bordered dataTable']//tbody//tr[@role='row']/td[position() = 2]")))]
    driver.quit()
    proxies = []
    for i in range(0, len(ips)):
        proxies.append(ips[i]+':'+ports[i])
    # print(proxies)
    return proxies


# proxies = GetProxies()

# options.add_argument('--headless')
# options.add_argument("--disable-notifications")


os.system('clear')

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
Dir_Name = os.path.join(BASE_DIR, 'TVShowsPythonScraper')
xl_file_path = os.path.join(Dir_Name, 'Shows.xlsx')
min_rating = 7
allowed_channels = ['hbo', 'netflix', 'hulu_plus', 'starz', 'showtime', 'apple_plus'] 

options = webdriver.ChromeOptions()
driver = webdriver.Chrome('chromedriver', options=options)
# driver = webdriver.Chrome('chromedriver.exe', options=options)
test_condition = 1
not_found_seasons = []


def ReadExcel():
    wb_obj = xl.load_workbook(xl_file_path)
    sheet_obj = wb_obj.active
    last_row_index_with_data = GetLastRowIndexWithData(sheet_obj) 
    # last_row_index_with_date = last row in the excel file with date 
    ratings_from_xl_file = []
    titles_from_xl_file = []
    mod_date_from_xl_file = []
    downloaded_seasons_from_xl_file = []
    total_seasons_from_xl_file = []

    for i in range(2, last_row_index_with_data+1):
        titles_from_xl_file.append(sheet_obj.cell(row=i, column=1).value)
        ratings_from_xl_file.append(sheet_obj.cell(row=i, column=2).value)
        mod_date_from_xl_file.append(sheet_obj.cell(row=i, column=4).value)
        downloaded_seasons_from_xl_file.append(sheet_obj.cell(row=i, column=5).value)
        total_seasons_from_xl_file.append(sheet_obj.cell(row=i, column=6).value)

    dictionary = { 
        'titles_from_xl_file':titles_from_xl_file, 
        'ratings_from_xl_file': ratings_from_xl_file, 
        'mod_date_from_xl_file': mod_date_from_xl_file,
        'downloaded_seasons_from_xl_file': downloaded_seasons_from_xl_file,
        'total_seasons_from_xl_file': total_seasons_from_xl_file
    }

    return dictionary

def GetLastRowIndexWithData(sheet_obj):
    number_of_rows = sheet_obj.max_row
    last_row_index_with_data = 0
    while True:
        try:
            # print(sheet_obj.cell(number_of_rows, 3).value)
            if sheet_obj.cell(row=number_of_rows, column=2).value != None:
                last_row_index_with_data = number_of_rows
                break
            elif number_of_rows == 1:
                return 1
                # last_row_index_with_data = number_of_rows
                # break
            else:
                number_of_rows -= 1
        except:
            print('something went wrong while reading from excel file')
    return last_row_index_with_data

def GetUnWantedTitles():
    is_last_page = False
    unwanted_titles = []
    
    # my_user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.111 Safari/537.36'
    counter = 0
    # while not is_last_page:
    while counter < test_condition:
        offset = counter*50
        counter += 1
        # url = f'https://reelgood.com/tv?offset={offset}'
        url = f'https://reelgood.com/tv/origin/america?filter-genre=6&filter-genre=39&filter-genre=15&filter-genre=16&filter-genre=18&filter-genre=37&offset={offset}'
        driver.get(url)
        # table_layout = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By., "myDynamicElement"))
        try:
            popup_X = driver.find_element_by_xpath('//div[@data-type="x"]')
            popup_X.click()
        except:
            pass
        
        try:
            table_layout = driver.find_element_by_xpath('//button[@title="Switch to table layout"]')
            table_layout.click()
        except:
            pass
        time.sleep(2)
        try:
            html = driver.page_source
            soup = BeautifulSoup(html, features='html.parser')
            table = soup.find('table', attrs={'class':'css-1179hly'})
        except:
            pass
        data = []
        if table: # checking if the table exists in the current page else the page will be last page 
            print(f'unwanted-titles being scraped at page # {counter}')
            # tbody = table.find('tbody')
            rows = soup.find_all('tr', attrs={'class':'css-gfsdx9'})
            for row in rows:
                td = row.find_all('td')
                td = [e.text.strip() for e in td]
                data.append([e for e in td if e])
            #  getting unwanted titles
            for row in data:
                unwanted_titles.append(row[0])# appending unwanted titles from website source
            # tds = soup.find_all('td', class_ = lambda value: value == 'css-1vuzpp2')
        else:
            is_last_page = True
    return unwanted_titles

def Process():
    is_last_page = False
    titles = []
    ratings = []
    number_of_seasons = []
    __available_on = []
    available_on = []
    unwanted_titles = GetUnWantedTitles()
    # print(unwanted_titles)
    # my_user_agent = 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.111 Safari/537.36'
    counter = 0
    # while not is_last_page:
    while counter < test_condition:
        offset = counter*50
        counter += 1
        # url = f'https://reelgood.com/tv?offset={offset}'
        url = f'https://reelgood.com/tv?filter-genre=5&filter-genre=3&filter-genre=13&filter-genre=26&filter-imdb_start=7&filter-rg_start=75&filter-year_start=2000&offset={offset}'
        driver.get(url)
        time.sleep(2)
        try:
            popup_X = driver.find_element_by_xpath('//div[@data-type="x"]')
            popup_X.click()# closing the popup
        except:
            pass
        try:
            table_layout = driver.find_element_by_xpath('//button[@title="Switch to table layout"]')
            table_layout.click()# changing the layout to tableview
        except:
            pass
        try:
            html = driver.page_source
            soup = BeautifulSoup(html, features='html.parser')
            table = soup.find('table', attrs={'class':'css-1179hly'})
        except:
            pass
        data = []
        # print(table)
        if table: # checking if the table exists in the current page else the page will be last page 
            print(f'general titles, ratings and channels are being scrapeed currently at page # {counter}')
            
            rows = soup.find_all('tr', attrs={'class':'css-gfsdx9'})
            for row in rows:
                td = row.find_all('td')
                row.find_all('')
                td = [e.text.strip() for e in td]
                
                data.append([e for e in td if e])
            #  getting titles and ratings
            #             
            for row in data:
                url = 'https://reelgood.com/show/' + row[0].lower().replace(' ', '-')+'-' + row[2]
                number_of_seasons.append(getNumberOfSeasons(url))
                if row[4].split('/')[0] == 'N': # if a rating is unavailable means N/A
                    rating = 0
                    ratings.append(rating)
                    titles.append(row[0]) 
                else:
                    rating = float(row[4].split('/')[0])
                    titles.append(row[0])# appending titles from website source
                    ratings.append(rating)# appending ratings from website source
                    # getting logos
            tds = soup.find_all('td', class_ = lambda value: value == 'css-1vuzpp2')
            
            for i, td in enumerate(tds):
                images_in_current_row = []
                images_in_current_row = td.find_all('img')
                logos = []
                for logo in images_in_current_row:
                    logos.append(logo['alt'])
                __available_on.append(logos)
                arr = []
                arr = __available_on[i] # populating arr with current row logo images. 
                temp = ''
                for a in arr:
                    temp += a
                    temp += ','
                available_on.append(temp[:-1]) # removing the last comma
           
        else:
            is_last_page = True
    
    dictionary = {'titles':titles, 'ratings': ratings, 'available_on': available_on, 'unwanted_titles': unwanted_titles, 'number_of_seasons': number_of_seasons}
    WriteToExcel(dictionary)

def getNumberOfSeasons(url):
    driver.execute_script(f'window.open("{url}");')
    driver.switch_to.window(driver.window_handles[1])
    total_seasons = len(driver.find_elements_by_xpath('//*[@id="app_mountpoint"]/div[3]/main/div/section/div[3]/div[2]/div/div[4]/div/div[2]/div[1]/div/*'))
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    return 0
    # for url in links:
    #     print(url,'this is url====================')
    #     driver.execute_script(f"window.open('https://reelgood.com{url}');")
    #     driver.switch_to.window(driver.window_handles[1])
    #     total_seasons = driver.find_element_by_xpath('//*[@id="app_mountpoint"]/div[3]/main/div/section/div[3]/div[2]/div/div[4]/div/div[2]/div[1]/div/*')
    #     number_of_seasons.append(total_seasons)
    #     print(total_seasons, 'this is total seasons')
    #     driver.close()
    #     driver.switch_to.window(driver.window_handles[0])
    # return number_of_seasons

def WriteToExcel(dictionary):
    number_of_seasons = dictionary['number_of_seasons']
    dictionary = ApplyFilter(dictionary)
    titles = dictionary['allowed_titles']
    ratings = dictionary['allowed_ratings']
    available_on = dictionary['allowed_available_on']
    # number_of_seasons = dictionary['number_of_seasons']
    
    wb = xl.load_workbook(xl_file_path)
    sheet = wb['Sheet1']
    sheet = wb.active
    ResetExcelSheetColors(sheet)
    dictionary_from_file = ReadExcel()
    titles_from_xl_file = dictionary_from_file['titles_from_xl_file']
    ratings_from_xl_file = dictionary_from_file['ratings_from_xl_file']
    downloaded_seasons_from_xl_file = dictionary_from_file['downloaded_seasons_from_xl_file']
    # mod_date_from_xl_file = dictionary_from_file['mod_date_from_xl_file']
    #  creting headings ----------------------------------
    sheet.cell(row=1, column=1).value = "Titles"
    sheet.cell(row=1, column=2).value = "Ratings "
    sheet.cell(row=1, column=3).value = "Available On"
    sheet.cell(row=1, column=4).value = "Download Date"
    sheet.cell(row=1, column=5).value = "Downloaded"
    sheet.cell(row=1, column=6).value = "# of Aired Seasons"
    # ----------------------------------------------------
    # print(f'titles from website length = {len(titles)}')
    # print(f'titles from excel length = {len(titles_from_xl_file)}')
    print('Please wait Excel file is being filled...')
    titles_left = []
    ratings_left = []
    available_on_left = []
    number_of_seasons_left = []
    if GetLastRowIndexWithData(sheet) == 1:
        print('Populating excel file for the first time.')
        for i, title in enumerate(titles):
            if ratings[i] >= min_rating:
                sheet.cell(row=i+2, column=1).value = title             # row[0] gets the title of TV Show
                sheet.cell(row=i+2, column=2).value = ratings[i]        # ratings gets the ratings of TV Show
                sheet.cell(row=i+2, column=3).value = available_on[i]
                sheet.cell(row=i+2, column=4).value = '-'
                sheet.cell(row=i+2, column=5).value = 'No'
                sheet.cell(row=i+2, column=6).value = number_of_seasons[i]
        wb.save(xl_file_path)
    elif min(len(titles), len(titles_from_xl_file)) == len(titles_from_xl_file):
        for i, title in enumerate(titles_from_xl_file):
            if titles_from_xl_file[i] != titles[i]: # checking if any titles has changed at website end.
                # sheet.cell(row=i+2, column=1).value = datetime.datetime.now().date()
                # sheet.cell(row=i+2, column=1).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
                sheet.cell(row=i+2, column=2).value = ratings[i] # ratings gets the ratings of TV Show
                # sheet.cell(row=i+2, column=3).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
                sheet.cell(row=i+2, column=3).value = available_on[i]
                # sheet.cell(row=i+2, column=4).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
            if ratings_from_xl_file[i] != ratings[i]: # checking if any rating has changed at website end.
                sheet.cell(row=i+2, column=2).value = ratings[i]
                sheet.cell(row=i+2, column=2).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
        wb.save(xl_file_path)

        low = len(titles_from_xl_file)
        high = len(titles)
        for i in range(low, high):
            titles_left.append(titles[i])
            ratings_left.append(ratings[i])
            available_on_left.append(available_on[i])
            number_of_seasons_left.append(number_of_seasons[i])
        # adding new entries at the end of the xl file
        for i in range(high-low):
            # sheet.cell(row=low+i, column=1).value = datetime.datetime.now().date()
            sheet.cell(row=low+i, column=1).value = titles_left[i]
            sheet.cell(row=low+i, column=2).value = ratings_left[i]
            sheet.cell(row=low+i, column=3).value = available_on_left[i]
            sheet.cell(row=low+i, column=4).value = '-'
            sheet.cell(row=low+i, column=5).value = 'No'
            sheet.cell(row=low+i, column=6).value = available_on_left[i]
        wb.save(xl_file_path)
    else:
        for i, title in enumerate(titles):
            if titles_from_xl_file[i] != titles[i]: # checking if any titles has changed at website end.
                # sheet.cell(row=i+2, column=1).value = datetime.datetime.now().date()
                # sheet.cell(row=i+2, column=1).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
                sheet.cell(row=i+2, column=1).value = titles[i]
                sheet.cell(row=i+2, column=1).fill = PatternFill(start_color='99FF99', end_color='99FF99',fill_type='solid')
                sheet.cell(row=i+2, column=2).value = ratings[i] # ratings gets the ratings of TV Show
                # sheet.cell(row=i+2, column=3).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
                sheet.cell(row=i+2, column=3).value = available_on[i]
                # sheet.cell(row=i+2, column=4).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
            if ratings_from_xl_file[i] != ratings[i]: # checking if any rating has changed at website end.
                # sheet.cell(row=i+2, column=1).value = datetime.datetime.now().date()
                # sheet.cell(row=i+2, column=1).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
                sheet.cell(row=i+2, column=2).value = ratings[i]
                sheet.cell(row=i+2, column=2).fill = PatternFill(start_color='FFFFB7', end_color='FFFFB7',fill_type='solid')
        wb.save(xl_file_path)
    wb.save(xl_file_path)
    # again reading excel file
    dictionary_from_file = ReadExcel()
    titles_from_xl_file = dictionary_from_file['titles_from_xl_file']
    number_of_seasons = dictionary_from_file['number_of_seasons']
    # CreateDirsFromListOfTitlesInExcelFile(titles_from_xl_file)

    options = webdriver.ChromeOptions()
    # print(random.choice(proxies), '-----------------------', len(proxies))
    # found = False
    # Pass the argument 1 to allow and 2 to block
    # options.setExperimentalOption("excludeSwitches",
    # array.asList("disable-popup-blocking"));
    options.add_argument(f"user-data-dir={Dir_Name}/profile");
    options.add_experimental_option("prefs", { 
        "profile.default_content_setting_values.notifications": 2,
        "disable-popup-blocking": True,
    })
    driver = webdriver.Chrome('chromedriver', options=options)
    driver.get('https://www.1377x.to/')

    for i, title in enumerate(titles_from_xl_file):
        # print(title, '--------------------------tilte')
        SearchForAllSeasonTorrent(driver, str(title), i, number_of_seasons[i] )

def WriteDownloadedSeasonNumberInExcel(show_index, season_no):
    wb = xl.load_workbook(xl_file_path)
    sheet = wb['Sheet1']
    sheet = wb.active
    sheet.cell(row=i+1, column=5).value = season_no
    wb.save(xl_file_path)

def SearchForAllSeasonTorrent(driver, title, i, total_seasons):
    time.sleep(random.choice([13]))
    for index in range(1, total_seasons):
        try:
            print('Searching for torrent, --------')
            search_input = driver.find_element_by_xpath('//*[@id="autocomplete"]')
            search_input.clear()
            search_input.send_keys(title.lower()+f' season {index}')
            search_input.send_keys(Keys.ENTER)
            
            time.sleep(2)
            CloseAllTabsExceptFirst(driver)
            print('Enter pressed ------------------------------')
            time.sleep(10)
            # result_rows = driver.find_element_by_xpath('//table/tbody/tr')
            # print(result_rows, ' this is rows variable')
            driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div[1]/table/tbody/tr[1]/td[1]/a[2]').click() # this is while cliking the first searched result

            time.sleep(1)
            CloseAllTabsExceptFirst(driver)
            # clicking the first search result

            print('open the first link in the search results -------------------')
            time.sleep(20)

            driver.find_element_by_xpath('/html/body/main/div/div/div/div[2]/div[1]/ul[1]/li[1]/a').click() 
            # clicking the magnet link
            time.sleep(1)
            CloseAllTabsExceptFirst(driver)

            
            # driver.find_element_by_xpath("//a[contains(.,'Magnet Download')]").click()

            print('clicked magnet link ==========================')
            WriteDownloadedSeasonNumberInExcel(i, season_no=index)
            # time.sleep(10)        
            time.sleep(2)
            driver.get('https://www.1377x.to/')
        except:
            pass
    time.sleep(random.choice([4, 5, 0.6, 2.9]))

def CloseAllTabsExceptFirst(driver):
    driver_len = len(driver.window_handles) #fetching the Number of Opened tabs
    print("Length of Driver = ", driver_len)
    if driver_len > 1: # Will execute if more than 1 tabs found.
        for i in range(driver_len - 1, 0, -1):
            driver.switch_to.window(driver.window_handles[i]) #will close the last tab first.
            driver.close()
            print("Closed Tab No. ", i)
        driver.switch_to.window(driver.window_handles[0]) # Switching the driver focus to First tab.
    else:
        print("Found only Single tab.")

def CreateDirsFromListOfTitlesInExcelFile(titles_from_xl_file):
    print(f'Please wait Directories are being created')
    # os.chdir('/')
    root = Dir_Name
    # root ='.'
    characters = [':', '*', '?', '\\', '/', '|', '"', '>', '<']
    for title in titles_from_xl_file:
        for c in characters:
            if c in title:
                title = title.replace(c , ' ')
        path = f'{root}/TV Shows/{title}'
        # print(path)
        if not os.path.exists(path):
            os.makedirs(path)
    print(f'{len(titles_from_xl_file)} folders created')

def IsAllowed(allowed_channels, ith_available_on):
    for val in allowed_channels:
        if val in ith_available_on:
            return True
    return False

def ResetExcelSheetColors(sheet):
    no_fill = xl.styles.PatternFill(fill_type=None)
    for row in sheet:
        for cell in row:
            cell.fill = no_fill    

def ApplyFilter(dictionary):
    titles = dictionary['titles']
    ratings = dictionary['ratings']
    available_on = dictionary['available_on']
    unwanted_titles = dictionary['unwanted_titles']
    allowed_ratings = []
    allowed_titles = []
    allowed_available_on = []
    for i, title in enumerate(titles):
        if ratings[i] >= min_rating:
            if IsAllowed(allowed_channels, available_on[i]):
                allowed_titles.append(title)
                allowed_ratings.append(ratings[i])
                allowed_available_on.append(available_on[i])
    for i, title in enumerate(allowed_titles):
        for unwanted_title in unwanted_titles:
            if unwanted_title == title:
                print(f'{allowed_titles[i]} is removed ')
                allowed_titles.pop(i)
                allowed_ratings.pop(i)
                allowed_available_on.pop(i)
    return {'allowed_ratings':allowed_ratings, 'allowed_titles': allowed_titles, 'allowed_available_on':allowed_available_on}

def RotateProxies():
    options = webdriver.ChromeOptions()
    options.add_argument("start-maximized")
    options.add_experimental_option("excludeSwitches", ["enable-automation"])
    options.add_experimental_option('useAutomationExtension', False)
    # driver = webdriver.Chrome(chrome_options=options, executable_path=r'C:\WebDrivers\chromedriver.exe')
    driver = webdriver.Chrome('chromedriver', options=options)

    driver.get("https://sslproxies.org/")
    driver.execute_script("return arguments[0].scrollIntoView(true);", WebDriverWait(driver, 20).until(EC.visibility_of_element_located((By.XPATH, "//table[@class='table table-striped table-bordered dataTable']//th[contains(., 'IP Address')]"))))
    ips = [my_elem.get_attribute("innerHTML") for my_elem in WebDriverWait(driver, 5).until(EC.visibility_of_all_elements_located((By.XPATH, "//table[@class='table table-striped table-bordered dataTable']//tbody//tr[@role='row']/td[position() = 1]")))]
    ports = [my_elem.get_attribute("innerHTML") for my_elem in WebDriverWait(driver, 5).until(EC.visibility_of_all_elements_located((By.XPATH, "//table[@class='table table-striped table-bordered dataTable']//tbody//tr[@role='row']/td[position() = 2]")))]
    driver.quit()
    proxies = []
    for i in range(0, len(ips)):
        proxies.append(ips[i]+':'+ports[i])
    print(proxies)
    options = webdriver.ChromeOptions()
    options.add_argument('--proxy-server={}'.format(random.choice(proxies)))

    # for i in range(0, len(proxies)):
    #     try:
    #         print("Proxy selected: {}".format(proxies[i]))

    #         options.add_argument('--proxy-server={}'.format(proxies[i]))
    #         # driver = webdriver.Chrome(options=options, executable_path=r'C:\WebDrivers\chromedriver.exe')
    #         driver = webdriver.Chrome('chromedriver', options=options)
    #         driver.get("https://www.whatismyip.com/proxy-check/?iref=home")
    #         if "Proxy Type" in WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "p.card-text"))):
    #             break
    #     except Exception:
    #         driver.quit()

    print("Proxy Invoked")


# starting point .............................. 
if __name__ == "__main__":
    # RotateProxies()
    Process()
    # dictionary = ReadExcel()
    # titles = dictionary["titles_from_xl_file"]
    # CreateDirsFromListOfTitlesInExcelFile(titles)
    driver.close()
    print('done')
    # schedule.every(1).day.at("01:00").do(Process)
    # while 1:
    #     schedule.run_pending()
    #     time.sleep(1)
